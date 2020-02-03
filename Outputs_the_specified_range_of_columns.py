#获取一个execl文件的所有sheet名
import openpyxl
import xlrd,xlwt
def range_col(start,end,sheet_names):
    wb = openpyxl.load_workbook(start)
    sheets = wb.sheetnames      #获取所有的execl名并输出到一个数组中：sheets
    #定义起始行数
    m=1
    #新建execl文件
    end_file=xlwt.Workbook(end)
    sheet1 = end_file.add_sheet(sheet_names)

    #将一个execl文件中指定列内容写入新建execl文件
    for i in range(len(sheets)):
        sheet_name=sheets[i]         #依次输出数组中的各个元素值
        print(sheet_name)
        file=xlrd.open_workbook(start)
        sheet_file=file.sheet_by_name(sheet_name)
        nrows = sheet_file.nrows       #行数
        ncols = sheet_file.ncols       #列数
        for r in range(nrows):
            r_value = sheet_file.row_values(r)
            j=0
            for p in r_value:
                if p < 2:
                    sheet1.write(m,j,p) #循环将第四列的值输出到新文件中
                    j=j+1
            m=m+1
    end_file.save(end)
range_col('shell.xlsx','end.xls','sheet')