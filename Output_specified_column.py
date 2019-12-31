#获取一个execl文件的所有sheet名
import openpyxl
import xlrd,xlwt
wb = openpyxl.load_workbook('shell.xlsx')
sheets = wb.sheetnames      #获取所有的execl名并输出到一个数组中：sheets
print(sheets)
#定义起始行数
m=1
#新建execl文件
end_file=xlwt.Workbook('end_file.xls')
sheet1 = end_file.add_sheet(u'sheet1')

#将一个execl文件中多个sheet表中指定行输出到新建execl中
for i in range(len(sheets)):
    sheet_name=sheets[i]         #依次输出数组中的各个元素值
    print(sheet_name)
    file=xlrd.open_workbook('shell.xlsx')
    sheet_file=file.sheet_by_name(sheet_name)
    nrows = sheet_file.nrows       #行数
    ncols = sheet_file.ncols       #列数
    for r in range(nrows):
        r_value = sheet_file.row_values(r)
        c_value=r_value[4]       #获取第几行的指定列值输出
        sheet1.write(m,5,c_value) #循环将第四列的值输出到新文件中
        m=m+1
end_file.save('end_file.xls')